#!/usr/bin/env python3
"""Extrai DADOS_DE_EMPLACAMENTO.xlsx (04/08/2026) → JSONs para o Smart Dealer.

Saídas (no scratchpad):
  varejo.json    — vendas por grupo/loja/modelo, meses 1–7/2026 (todas as 9 CCYs)
  estoque.json   — estoque chão/trânsito por grupo/loja/modelo (todas as CCYs)
  metas.json     — carta de agosto por grupo
  share.json     — emplacamento das áreas da Nippon (Amparo + Ouro Fino), meses 1–7
  calendario.json— dias úteis de agosto/2026
"""
import json, re, unicodedata
from pathlib import Path
import openpyxl

SRC = Path('/Users/caiqueoliveira/Documents/MEUS PROJETOS/_Performance Concessionário/DADOS_DE_EMPLACAMENTO.xlsx')
OUT = Path(__file__).parent

AREAS_NIPPON = {'Amparo', 'Ouro Fino'}
MES_FECHADO = 7      # julho
MES_CORRENTE = 8     # agosto (carta)
ANO = 2026

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

def norm_loja(abremp: str) -> str:
    """'NIPPON MOTOS - BRAGANCA PAULISTA' → 'Bragança Paulista' etc."""
    s = str(abremp).strip()
    if ' - ' in s:
        s = s.split(' - ', 1)[1]
    else:
        # 'NIPPON MOTOS AMPARO' → tira o prefixo do grupo se repetir
        parts = s.split()
        s = ' '.join(parts[2:]) if s.upper().startswith('NIPPON MOTOS') and len(parts) > 2 else s
    s = s.strip().title()
    fixes = {'Braganca Paulista': 'Bragança Paulista', 'Braganca': 'Bragança Paulista'}
    return fixes.get(s, s)

# ---------- VAREJO+META ----------
ws = wb['VAREJO+META']
rows = list(ws.iter_rows(values_only=True))
hdr = next(i for i, r in enumerate(rows) if r[0] == 'GRUPO')
varejo = []
for r in rows[hdr + 1:]:
    if not r[0]:
        continue
    for m in range(1, 8):
        q = r[3 + m - 1]
        if isinstance(q, (int, float)) and q > 0:
            varejo.append({'grupo': str(r[0]), 'loja': norm_loja(r[1]), 'modelo': str(r[2]).strip(),
                           'mes': m, 'ano': ANO, 'quantidade': int(q)})

# ---------- ESTOQUE ----------
ws = wb['ESTOQUE']
rows = list(ws.iter_rows(values_only=True))
hdr = next(i for i, r in enumerate(rows) if r[0] == 'GRUPO')
estoque = []
for r in rows[hdr + 1:]:
    if not r[0]:
        continue
    # Mesma regra do Performance Concessionário: chão vem da coluna CHÃO;
    # trânsito = rede − chão (a coluna EM TRÂNSITO do arquivo é inconsistente
    # com REDE em várias linhas — o total confiável é REDE).
    chao = int(r[4]) if isinstance(r[4], (int, float)) else 0
    rede = int(r[5]) if isinstance(r[5], (int, float)) else 0
    transito = max(rede - chao, 0)
    estoque.append({'grupo': str(r[0]), 'loja': norm_loja(r[1]), 'modelo': str(r[2]).strip(),
                    'chao': chao, 'transito': transito})

# ---------- DIAS DE ESTOQUE ----------
ws = wb['DIAS DE ESTOQUE']
rows = list(ws.iter_rows(values_only=True))
hdr = next(i for i, r in enumerate(rows) if r[0] == 'GRUPO')
data_estoque = next((str(r[1]) for r in rows if r[0] and 'Data' in str(r[0])), None)
dias_estoque = []
for r in rows[hdr + 1:]:
    if not r[0]:
        continue
    v = r[3]
    dias = round(float(v), 1) if isinstance(v, (int, float)) else None  # '#NUM!' → None (sem giro)
    dias_estoque.append({'grupo': str(r[0]), 'loja': norm_loja(r[1]), 'modelo': str(r[2]).strip(), 'dias': dias})

# ---------- META MÊS ATUAL ----------
ws = wb['META MÊS ATUAL']
metas = {}
for r in list(ws.iter_rows(values_only=True))[1:]:
    if r[2] and isinstance(r[3], (int, float)):
        metas[str(r[2]).strip()] = int(r[3])

# ---------- EMPLACAMENTO (mercado das áreas da Nippon) ----------
ws = wb['EMPLACAMENTO']
rows = list(ws.iter_rows(values_only=True))
hdr = next(i for i, r in enumerate(rows) if r[0] == 'CNPJ')
share = []
for r in rows[hdr + 1:]:
    if not r[0] or str(r[2]) not in AREAS_NIPPON:
        continue
    meses = {str(m): int(r[5 + m - 1]) if isinstance(r[5 + m - 1], (int, float)) else 0 for m in range(1, 8)}
    if sum(meses.values()) == 0:
        continue
    share.append({'cnpj': str(r[0]), 'marca': str(r[1]), 'area': str(r[2]),
                  'cidade': str(r[3]).strip().title(), 'segmento': str(r[4]), 'meses': meses})

# ---------- CALENDÁRIO (dias úteis de agosto) ----------
ws = wb['CALENDÁRIO - WORKING DAYS']
cal_rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
# procurar estrutura: colunas com mês/dia — dump bruto pra inspecionar depois
cal_sample = cal_rows[:30]

json.dump(varejo, open(OUT / 'varejo.json', 'w'), ensure_ascii=False)
json.dump(estoque, open(OUT / 'estoque.json', 'w'), ensure_ascii=False)
json.dump(dias_estoque, open(OUT / 'dias_estoque.json', 'w'), ensure_ascii=False)
json.dump(metas, open(OUT / 'metas.json', 'w'), ensure_ascii=False, indent=1)
json.dump(share, open(OUT / 'share.json', 'w'), ensure_ascii=False)
json.dump({'data_estoque': data_estoque, 'cal_sample': [[str(c) if c is not None else None for c in r[:13]] for r in cal_sample]},
          open(OUT / 'meta_info.json', 'w'), ensure_ascii=False, indent=1)

# resumo
tot_por_grupo = {}
for v in varejo:
    if v['mes'] == MES_FECHADO:
        tot_por_grupo[v['grupo']] = tot_por_grupo.get(v['grupo'], 0) + v['quantidade']
print('Varejo julho por grupo:', dict(sorted(tot_por_grupo.items())))
print('Linhas: varejo', len(varejo), '| estoque', len(estoque), '| dias', len(dias_estoque), '| share', len(share))
print('Metas agosto:', metas)
print('Data estoque:', data_estoque)
nip_share_jul = sum(s['meses']['7'] for s in share)
nip_yam_jul = sum(s['meses']['7'] for s in share if s['marca'] == 'Yamaha')
print(f'Mercado áreas Nippon jul: {nip_share_jul} | Yamaha: {nip_yam_jul} ({100*nip_yam_jul/nip_share_jul:.1f}%)')
